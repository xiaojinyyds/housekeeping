#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""合同导入导出接口"""
from datetime import datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user_id, get_current_user_role
from app.models.business import CustomerLead, ServiceContract
from app.models.user import User
from app.schemas.response import ApiResponse
from app.utils.helpers import generate_uuid

router = APIRouter()

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="FF9F43", end_color="FF9F43", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def get_content_disposition(filename: str) -> str:
    encoded_filename = quote(filename)
    return f"attachment; filename*=UTF-8''{encoded_filename}"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                text = str(cell.value)
                size = len(text) + sum(1 for ch in text if "\u4e00" <= ch <= "\u9fff")
                max_length = max(max_length, size)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def parse_date_safe(value: Optional[str]):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except Exception:
        return None


def parse_datetime_safe(value: Optional[str]):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


@router.get("/import/template/contracts", summary="下载合同导入模板")
async def download_contract_template(user_role: str = Depends(get_current_user_role)):
    if user_role not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    wb = Workbook()
    ws = wb.active
    ws.title = "合同导入模板"

    headers = [
        "合同日期*(YYYY-MM-DD)", "线索ID*", "阿姨用户ID*", "客户姓名*", "客户电话*", "服务地址*", "订单类型*",
        "签单金额", "实收金额", "签约时间(YYYY-MM-DDTHH:mm:ss)", "上户时间(YYYY-MM-DDTHH:mm:ss)",
        "到期时间(YYYY-MM-DDTHH:mm:ss)", "合同状态", "签单员工ID(仅管理员可填)", "备注"
    ]
    ws.append(headers)
    style_header(ws)
    ws.append([
        datetime.now().strftime("%Y-%m-%d"), "lead-id-demo", "worker-user-id-demo", "王女士", "13800138000", "朝阳区某小区",
        "住家保姆", 9800, 9800, datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "", "", "pending_start", "", "首次签约"
    ])

    ws_help = wb.create_sheet("填写说明")
    ws_help["A1"] = "【填写说明】"
    ws_help["A1"].font = Font(bold=True, color="FF0000", size=14)
    ws_help["A3"] = "1. 带 * 字段为必填"
    ws_help["A4"] = "2. 合同状态可选：pending_start / serving / paused / completed / terminated / refunded"
    ws_help["A5"] = "3. 员工导入时，签单员工ID会自动归属当前登录账号"
    ws_help["A6"] = "4. 线索ID必须存在，且员工只能导入自己负责线索"
    ws_help["A7"] = "5. 合同编号由系统自动生成"
    ws_help.column_dimensions["A"].width = 88

    auto_column_width(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition("合同导入模板.xlsx")}
    )


@router.post("/import/contracts", summary="批量导入合同")
async def import_contracts(
    file: UploadFile = File(..., description="Excel文件"),
    current_user_id: str = Depends(get_current_user_id),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    if user_role not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传Excel文件")

    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.worksheets[0]
        results = {"success": 0, "failed": 0, "errors": []}
        valid_statuses = {"pending_start", "serving", "paused", "completed", "terminated", "refunded"}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:
                continue
            first_cell = str(row[0]).strip() if row[0] else ""
            if first_cell.startswith("【") or first_cell.endswith("*") or first_cell == "合同日期":
                continue

            try:
                contract_date = parse_date_safe(str(row[0]).strip()) if row[0] else None
                lead_id = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                worker_user_id = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                customer_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                customer_phone = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                service_address = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                service_type = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                contract_amount_raw = row[7] if len(row) > 7 else None
                actual_received_raw = row[8] if len(row) > 8 else None
                sign_date = parse_datetime_safe(str(row[9]).strip()) if len(row) > 9 and row[9] else None
                start_date = parse_datetime_safe(str(row[10]).strip()) if len(row) > 10 and row[10] else None
                end_date = parse_datetime_safe(str(row[11]).strip()) if len(row) > 11 and row[11] else None
                status_value = str(row[12]).strip() if len(row) > 12 and row[12] else "pending_start"
                broker_staff_id_excel = str(row[13]).strip() if len(row) > 13 and row[13] else ""
                remark = str(row[14]).strip() if len(row) > 14 and row[14] else ""

                if not all([contract_date, lead_id, worker_user_id, customer_name, customer_phone, service_address, service_type]):
                    results["errors"].append(f"第{row_idx}行: 缺少必填字段")
                    results["failed"] += 1
                    continue

                lead = db.query(CustomerLead).filter(CustomerLead.id == lead_id).first()
                if not lead:
                    results["errors"].append(f"第{row_idx}行: 线索ID不存在")
                    results["failed"] += 1
                    continue
                if user_role == "staff" and lead.owner_staff_id != current_user_id:
                    results["errors"].append(f"第{row_idx}行: 不能导入非本人负责线索")
                    results["failed"] += 1
                    continue

                worker = db.query(User).filter(User.id == worker_user_id, User.role == "worker").first()
                if not worker:
                    results["errors"].append(f"第{row_idx}行: 阿姨用户ID不存在")
                    results["failed"] += 1
                    continue

                broker_staff_id = current_user_id
                if user_role == "admin" and broker_staff_id_excel:
                    staff = db.query(User).filter(
                        User.id == broker_staff_id_excel,
                        User.role == "staff",
                        User.status == "active"
                    ).first()
                    if staff:
                        broker_staff_id = staff.id

                contract_amount = None
                if contract_amount_raw not in [None, ""]:
                    try:
                        contract_amount = float(contract_amount_raw)
                    except Exception:
                        results["errors"].append(f"第{row_idx}行: 签单金额格式错误")
                        results["failed"] += 1
                        continue

                actual_received = None
                if actual_received_raw not in [None, ""]:
                    try:
                        actual_received = float(actual_received_raw)
                    except Exception:
                        results["errors"].append(f"第{row_idx}行: 实收金额格式错误")
                        results["failed"] += 1
                        continue

                if status_value not in valid_statuses:
                    results["errors"].append(f"第{row_idx}行: 合同状态无效")
                    results["failed"] += 1
                    continue

                contract = ServiceContract(
                    id=generate_uuid(),
                    contract_no=f"HT{datetime.now().strftime('%Y%m%d%H%M%S')}{str(row_idx).zfill(3)}",
                    lead_id=lead.id,
                    worker_user_id=worker.id,
                    broker_staff_id=broker_staff_id,
                    customer_name=customer_name,
                    customer_phone=customer_phone,
                    customer_source=lead.source,
                    service_address=service_address,
                    service_type=service_type,
                    contract_date=contract_date,
                    sign_date=sign_date,
                    start_date=start_date,
                    end_date=end_date,
                    status=status_value,
                    contract_amount=contract_amount,
                    actual_received=actual_received,
                    remark=remark or None
                )
                db.add(contract)
                lead.status = "signed"
                lead.latest_follow_up_at = datetime.now()
                results["success"] += 1
            except Exception as e:
                results["errors"].append(f"第{row_idx}行: {str(e)}")
                results["failed"] += 1

        db.commit()
        return ApiResponse.success(
            data=results,
            message=f"导入完成：成功{results['success']}条，失败{results['failed']}条"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入失败: {str(e)}"
        )
