#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""数据导入导出API"""
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, String as SAString
from datetime import datetime
from typing import Optional
from io import BytesIO
from urllib.parse import quote
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.core.database import get_db
from app.core.security import get_current_user_id, get_current_user_role, get_password_hash
from app.models.business import CustomerLead, ServiceContract
from app.models.user import User
from app.models.worker import WorkerProfile
from app.schemas.response import ApiResponse
from app.utils.helpers import generate_uuid

router = APIRouter()

# Excel 样式
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="FF9F43", end_color="FF9F43", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_content_disposition(filename: str) -> str:
    """生成 Content-Disposition header，支持中文文件名"""
    encoded_filename = quote(filename)
    return f"attachment; filename*=UTF-8''{encoded_filename}"


def style_header(ws, row=1):
    """设置表头样式"""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    # 中文字符算2个宽度
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def parse_date_safe(value: Optional[str]):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def parse_datetime_safe(value: Optional[str]):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


# ==================== 导出接口 ====================

@router.get("/export/workers", summary="导入阿姨")
async def export_workers(
    is_available: Optional[bool] = Query(None, description="是否可接单"),
    keyword: Optional[str] = Query(None, description="关键词(姓名/手机/身份证)"),
    address: Optional[str] = Query(None, description="????"),
    service_area: Optional[str] = Query(None, description="????"),
    job_type: Optional[str] = Query(None, description="????"),
    current_status: Optional[str] = Query(None, description="????"),
    min_age: Optional[int] = Query(None, ge=0, description="????"),
    max_age: Optional[int] = Query(None, ge=0, description="????"),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    query = db.query(WorkerProfile)
    if is_available is not None:
        query = query.filter(WorkerProfile.is_available == is_available)
    if keyword:
        query = query.filter(
            (WorkerProfile.real_name.like(f"%{keyword}%")) |
            (WorkerProfile.phone.like(f"%{keyword}%")) |
            (WorkerProfile.id_card.like(f"%{keyword}%"))
        )
    if address:
        query = query.filter(WorkerProfile.address.like(f"%{address}%"))
    if service_area:
        query = query.filter(func.cast(WorkerProfile.service_areas, SAString).like(f"%{service_area}%"))
    if job_type:
        query = query.filter(func.cast(WorkerProfile.job_types, SAString).like(f"%{job_type}%"))
    if current_status:
        query = query.filter(WorkerProfile.current_status == current_status)
    if min_age is not None:
        query = query.filter(WorkerProfile.age >= min_age)
    if max_age is not None:
        query = query.filter(WorkerProfile.age <= max_age)

    workers = query.order_by(WorkerProfile.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "阿姨档案"
    headers = [
        "创建时间", "姓名", "手机号", "身份证号", "性别", "年龄", "从业年限", "接单类型",
        "技能标签", "居住地址", "服务区域", "微信号", "期望薪资", "当前状态", "可接单", "个人简介"
    ]
    ws.append(headers)
    style_header(ws)

    gender_map = {"male": "男", "female": "女"}
    status_map = {
        "available": "可接单",
        "on_job": "上户中",
        "paused": "暂停",
        "blacklisted": "黑名单",
        "inactive": "停用"
    }

    for item in workers:
        ws.append([
            item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "",
            item.real_name or "",
            item.phone or "",
            item.id_card or "",
            gender_map.get(item.gender, item.gender or ""),
            item.age or "",
            item.experience_years or "",
            ",".join(item.job_types or []),
            ",".join(item.skills or []),
            item.address or "",
            ",".join(item.service_areas or []),
            item.wechat or "",
            float(item.expected_salary) if item.expected_salary is not None else "",
            status_map.get(item.current_status, item.current_status or ""),
            "是" if item.is_available else "否",
            item.introduction or ""
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
    auto_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"阿姨档案_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition(filename)}
    )


@router.get("/export/users", summary="导出用户列表")
async def export_users(
    role: Optional[str] = Query(None, description="筛选角色"),
    status_filter: Optional[str] = Query(None, alias="status", description="筛选状态"),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """导出用户列表为Excel"""
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")
    
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    if status_filter:
        query = query.filter(User.status == status_filter)
    
    users = query.order_by(User.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "用户列表"
    
    headers = ["昵称", "邮箱", "手机号", "角色", "状态", "注册时间", "最后登录"]
    ws.append(headers)
    style_header(ws)
    
    role_map = {"user": "普通用户", "worker": "家政阿姨", "admin": "管理员"}
    status_map = {"active": "正常", "disabled": "禁用"}
    
    for u in users:
        ws.append([
            u.nickname or "-",
            u.email,
            u.phone or "-",
            role_map.get(u.role, u.role),
            status_map.get(u.status, u.status),
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            u.last_login_at.strftime("%Y-%m-%d %H:%M") if u.last_login_at else "-"
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"用户列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition(filename)}
    )


@router.get("/export/orders", summary="导出订单数据")
async def export_orders(
    status_filter: Optional[str] = Query(None, alias="status", description="订单状态筛选"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """
    导出订单数据为Excel
    
    支持筛选：
    - status: 订单状态
    - start_date/end_date: 日期范围
    """
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")
    
    from app.models.appointment import Appointment
    from sqlalchemy.orm import aliased
    
    # 创建别名用于关联查询
    UserAlias = aliased(User)
    WorkerUser = aliased(User)
    
    query = db.query(Appointment)
    
    if status_filter:
        query = query.filter(Appointment.status == status_filter)
    
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Appointment.created_at >= start)
        except:
            pass
    
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            end = end.replace(hour=23, minute=59, second=59)
            query = query.filter(Appointment.created_at <= end)
        except:
            pass
    
    orders = query.order_by(Appointment.created_at.desc()).all()
    
    # 获取用户和阿姨信息
    user_ids = set()
    worker_ids = set()
    for o in orders:
        if o.user_id:
            user_ids.add(o.user_id)
        if o.worker_id:
            worker_ids.add(o.worker_id)
    
    users_map = {}
    workers_map = {}
    
    if user_ids:
        users = db.query(User).filter(User.id.in_(user_ids)).all()
        users_map = {u.id: u for u in users}
    
    if worker_ids:
        workers = db.query(WorkerProfile).filter(WorkerProfile.user_id.in_(worker_ids)).all()
        workers_map = {w.user_id: w for w in workers}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "订单数据"
    
    headers = ["订单号", "用户", "用户电话", "阿姨", "阿姨电话", "服务项目", 
               "预约日期", "时间段", "时长(h)", "单价", "总价", "服务地址",
               "联系人", "联系电话", "状态", "创建时间"]
    ws.append(headers)
    style_header(ws)
    
    status_map = {
        "pending": "待接单",
        "accepted": "已接单", 
        "rejected": "已拒绝",
        "in_progress": "服务中",
        "completed": "已完成",
        "reviewed": "已评价",
        "cancelled": "已取消"
    }
    
    for o in orders:
        user = users_map.get(o.user_id)
        worker = workers_map.get(o.worker_id)
        
        ws.append([
            o.order_no,
            user.nickname or user.email.split('@')[0] if user else "-",
            user.phone if user else "-",
            worker.real_name if worker else "-",
            worker.phone if worker else "-",
            o.service_name or "-",
            o.appointment_date.strftime("%Y-%m-%d") if o.appointment_date else "-",
            o.time_slot_name or "-",
            o.duration_hours or 0,
            float(o.unit_price) if o.unit_price else 0,
            float(o.total_price) if o.total_price else 0,
            o.address or "-",
            o.contact_name or "-",
            o.contact_phone or "-",
            status_map.get(o.status, o.status),
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else ""
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"订单数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition(filename)}
    )


@router.get("/export/leads", summary="导出客户线索")
async def export_leads(
    owner_staff_id: Optional[str] = Query(None, description="负责人ID"),
    status_filter: Optional[str] = Query(None, alias="status", description="线索状态"),
    lead_category: Optional[str] = Query(None, description="线索类别 A/B"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """导出客户线索（仅管理员）"""
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    query = db.query(CustomerLead)
    if owner_staff_id:
        query = query.filter(CustomerLead.owner_staff_id == owner_staff_id)
    if status_filter:
        query = query.filter(CustomerLead.status == status_filter)
    if lead_category:
        query = query.filter(CustomerLead.lead_category == lead_category)

    start = parse_date_safe(start_date)
    end = parse_date_safe(end_date)
    if start:
        query = query.filter(CustomerLead.lead_date >= start)
    if end:
        query = query.filter(CustomerLead.lead_date <= end)

    leads = query.order_by(CustomerLead.created_at.desc()).all()

    staff_ids = list({lead.owner_staff_id for lead in leads if lead.owner_staff_id})
    staff_map = {}
    if staff_ids:
        staffs = db.query(User).filter(User.id.in_(staff_ids)).all()
        staff_map = {s.id: s for s in staffs}

    wb = Workbook()
    ws = wb.active
    ws.title = "客户线索"

    headers = [
        "线索编号", "线索日期", "客户称呼", "联系电话", "线索类别", "线索状态", "负责人",
        "订单类型", "客户来源", "来源补充", "线索名称", "需求地址", "预算", "备注", "创建时间"
    ]
    ws.append(headers)
    style_header(ws)

    for lead in leads:
        owner = staff_map.get(lead.owner_staff_id)
        owner_name = (owner.nickname or owner.real_name or owner.email) if owner else ""
        ws.append([
            lead.lead_no or "",
            lead.lead_date.strftime("%Y-%m-%d") if lead.lead_date else "",
            lead.customer_name or "",
            lead.phone or "",
            lead.lead_category or "B",
            lead.status or "new",
            owner_name,
            lead.service_type or "",
            lead.source or "",
            lead.source_detail or "",
            lead.lead_name or "",
            lead.demand_address or "",
            float(lead.budget) if lead.budget else "",
            lead.remark or "",
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else ""
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
    auto_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"客户线索_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition(filename)}
    )


@router.get("/export/contracts", summary="导出合同数据")
async def export_contracts(
    broker_staff_id: Optional[str] = Query(None, description="签单员工ID"),
    status_filter: Optional[str] = Query(None, alias="status", description="合同状态"),
    service_type: Optional[str] = Query(None, description="订单类型"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """导出合同数据（仅管理员）"""
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    query = db.query(ServiceContract)
    if broker_staff_id:
        query = query.filter(ServiceContract.broker_staff_id == broker_staff_id)
    if status_filter:
        query = query.filter(ServiceContract.status == status_filter)
    if service_type:
        query = query.filter(ServiceContract.service_type == service_type)

    start = parse_date_safe(start_date)
    end = parse_date_safe(end_date)
    if start:
        query = query.filter(ServiceContract.contract_date >= start)
    if end:
        query = query.filter(ServiceContract.contract_date <= end)

    contracts = query.order_by(ServiceContract.contract_date.desc(), ServiceContract.created_at.desc()).all()

    user_ids = list({
        uid for item in contracts for uid in [item.worker_user_id, item.broker_staff_id, item.customer_user_id] if uid
    })
    user_map = {}
    if user_ids:
        users = db.query(User).filter(User.id.in_(user_ids)).all()
        user_map = {u.id: u for u in users}

    lead_ids = list({item.lead_id for item in contracts if item.lead_id})
    lead_map = {}
    if lead_ids:
        leads = db.query(CustomerLead).filter(CustomerLead.id.in_(lead_ids)).all()
        lead_map = {item.id: item for item in leads}

    wb = Workbook()
    ws = wb.active
    ws.title = "合同数据"
    headers = [
        "合同编号", "合同日期", "签约时间", "线索ID", "线索编号", "客户姓名", "客户电话", "订单类型", "服务地址",
        "合同状态", "签单金额", "实收金额", "签单员工ID", "签单员工", "阿姨ID", "阿姨", "创建时间", "备注"
    ]
    ws.append(headers)
    style_header(ws)

    for item in contracts:
        broker = user_map.get(item.broker_staff_id)
        worker = user_map.get(item.worker_user_id)
        lead = lead_map.get(item.lead_id)
        ws.append([
            item.contract_no or "",
            item.contract_date.strftime("%Y-%m-%d") if item.contract_date else "",
            item.sign_date.strftime("%Y-%m-%d %H:%M:%S") if item.sign_date else "",
            item.lead_id or "",
            lead.lead_no if lead else "",
            item.customer_name or "",
            item.customer_phone or "",
            item.service_type or "",
            item.service_address or "",
            item.status or "",
            float(item.contract_amount) if item.contract_amount is not None else "",
            float(item.actual_received) if item.actual_received is not None else "",
            item.broker_staff_id or "",
            (broker.nickname or getattr(broker, "real_name", None) or broker.email) if broker else "",
            item.worker_user_id or "",
            (worker.nickname or getattr(worker, "real_name", None) or worker.email) if worker else "",
            item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else "",
            item.remark or ""
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
    auto_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"合同数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition(filename)}
    )


@router.get("/export/statistics", summary="导出统计报告")
async def export_statistics(
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """导出平台统计报告为Excel"""
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")
    
    from app.models.worker import WorkerApplication
    
    wb = Workbook()
    ws = wb.active
    ws.title = "平台统计报告"
    
    # 报告标题
    ws.merge_cells('A1:C1')
    ws['A1'] = f"家政服务平台统计报告 - {datetime.now().strftime('%Y年%m月%d日')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    row = 3
    
    # 用户统计
    ws[f'A{row}'] = "【用户统计】"
    ws[f'A{row}'].font = Font(bold=True, color="FF9F43")
    row += 1
    
    total_users = db.query(func.count(User.id)).scalar()
    active_users = db.query(func.count(User.id)).filter(User.status == 'active').scalar()
    user_count = db.query(func.count(User.id)).filter(User.role == 'user').scalar()
    worker_count = db.query(func.count(User.id)).filter(User.role == 'worker').scalar()
    admin_count = db.query(func.count(User.id)).filter(User.role == 'admin').scalar()
    
    stats = [
        ("总用户数", total_users),
        ("活跃用户", active_users),
        ("普通用户", user_count),
        ("家政阿姨", worker_count),
        ("管理员", admin_count),
    ]
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # 阿姨统计
    ws[f'A{row}'] = "【阿姨统计】"
    ws[f'A{row}'].font = Font(bold=True, color="FF9F43")
    row += 1
    
    total_workers = db.query(func.count(WorkerProfile.id)).scalar()
    available_workers = db.query(func.count(WorkerProfile.id)).filter(WorkerProfile.is_available == True).scalar()
    recommended_workers = db.query(func.count(WorkerProfile.id)).filter(WorkerProfile.is_recommended == True).scalar()
    avg_rating = db.query(func.avg(WorkerProfile.rating)).scalar() or 0
    
    stats = [
        ("阿姨总数", total_workers),
        ("接单中", available_workers),
        ("首页推荐", recommended_workers),
        ("平均评分", f"{float(avg_rating):.2f}"),
    ]
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # 申请统计
    ws[f'A{row}'] = "【申请统计】"
    ws[f'A{row}'].font = Font(bold=True, color="FF9F43")
    row += 1
    
    pending = db.query(func.count(WorkerApplication.id)).filter(WorkerApplication.status == 'pending').scalar()
    approved = db.query(func.count(WorkerApplication.id)).filter(WorkerApplication.status == 'approved').scalar()
    rejected = db.query(func.count(WorkerApplication.id)).filter(WorkerApplication.status == 'rejected').scalar()
    
    stats = [
        ("待审核", pending),
        ("已通过", approved),
        ("已拒绝", rejected),
        ("总申请", pending + approved + rejected),
    ]
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"平台统计报告_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition(filename)}
    )


# ==================== 导入模板 ====================

@router.get("/import/template/workers", summary="下载阿姨导入模板")
async def download_worker_template(
    user_role: str = Depends(get_current_user_role)
):
    if user_role not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    wb = Workbook()
    ws = wb.active
    ws.title = "阿姨档案"

    headers = [
        "姓名*", "手机号*", "身份证号*", "性别*(男/女)", "年龄*", "从业年限*", "接单类型",
        "技能标签*", "居住地址*", "微信号", "期望薪资", "当前状态", "可接单*(是/否)", "个人简介*"
    ]
    ws.append(headers)
    style_header(ws)

    ws.append([
        "张三", "13800138001", "110101199001011234", "女", 42, 8, "住家保姆,月嫂",
        "做饭,保洁,带娃", "北京市朝阳区", "zhangyi", 8500,
        "available", "是", "认真负责，保洁经验丰富"
    ])

    ws_help = wb.create_sheet("填写说明")
    ws_help["A1"] = "阿姨档案导入说明"
    ws_help["A1"].font = Font(bold=True, color="FF0000", size=14)
    ws_help["A3"] = "1. * 为必填项"
    ws_help["A4"] = "2. 性别填写：男/女"
    ws_help["A5"] = "3. 当前状态填写：available / on_job / paused / blacklisted / inactive"
    ws_help["A6"] = "4. 接单类型/技能标签用英文逗号分隔"
    ws_help["A7"] = "5. 可同时导入多条记录"
    ws_help["A8"] = "6. 导入前请查看填写说明"
    ws_help.column_dimensions["A"].width = 88

    auto_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition("阿姨档案导入模板.xlsx")}
    )


@router.get("/import/template/users", summary="下载用户导入模板")
async def download_user_template(
    user_role: str = Depends(get_current_user_role)
):
    """下载用户批量导入的Excel模板"""
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    
    headers = ["邮箱*", "手机号", "昵称"]
    ws.append(headers)
    style_header(ws)
    
    ws.append(["user1@example.com", "13900139001", "用户1"])
    ws.append(["user2@example.com", "13900139002", "用户2"])
    
    # 添加说明到单独的sheet
    ws_help = wb.create_sheet("填写说明")
    ws_help['A1'] = "【填写说明】"
    ws_help['A1'].font = Font(bold=True, color="FF0000", size=14)
    ws_help['A3'] = "1. 邮箱为必填项，不能重复"
    ws_help['A4'] = "2. 手机号可选，但不能重复"
    ws_help['A5'] = "3. 导入后系统会自动创建账号，默认密码为123456"
    ws_help['A6'] = "4. 请在【用户导入模板】sheet中填写数据，从第2行开始"
    ws_help.column_dimensions['A'].width = 60
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition("用户导入模板.xlsx")}
    )


@router.get("/import/template/leads", summary="下载线索导入模板")
async def download_lead_template(
    user_role: str = Depends(get_current_user_role)
):
    """下载线索导入模板（管理员/员工）"""
    if user_role not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    wb = Workbook()
    ws = wb.active
    ws.title = "线索导入模板"

    headers = [
        "线索日期*(YYYY-MM-DD)", "客户称呼*", "联系电话*", "线索类别*(A/B)",
        "订单类型*", "客户来源", "来源补充", "线索名称", "需求地址", "需求详情", "预算", "备注", "负责人ID(管理员可填)"
    ]
    ws.append(headers)
    style_header(ws)

    ws.append([
        datetime.now().strftime("%Y-%m-%d"), "王女士", "13800138000", "A",
        "住家保姆", "转介绍", "张老师", "高净值客户", "朝阳区XX小区", "希望尽快到岗", 9000, "重点跟进", ""
    ])

    ws_help = wb.create_sheet("填写说明")
    ws_help["A1"] = "【填写说明】"
    ws_help["A1"].font = Font(bold=True, color="FF0000", size=14)
    ws_help["A3"] = "1. 带*字段为必填"
    ws_help["A4"] = "2. 线索类别仅支持 A 或 B"
    ws_help["A5"] = "3. 管理员可填写负责人ID；员工导入时会自动归属自己"
    ws_help["A6"] = "4. 线索日期为空时默认当天"
    ws_help["A7"] = "5. 线索状态默认 new（新线索）"
    ws_help.column_dimensions["A"].width = 80

    auto_column_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": get_content_disposition("线索导入模板.xlsx")}
    )


# ==================== 导入接口 ====================

@router.post("/import/workers", summary="导入阿姨")
async def import_workers(
    file: UploadFile = File(..., description="Excel文件"),
    current_user_id: str = Depends(get_current_user_id),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    if user_role not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传 Excel 文件")

    def split_multi_value(value):
        if value in [None, ""]:
            return []
        return [item.strip() for item in str(value).replace("，", ",").split(",") if item.strip()]

    def parse_gender(value):
        raw = str(value or "").strip()
        if raw in ["女", "female", "Female"]:
            return "female"
        if raw in ["男", "male", "Male"]:
            return "male"
        return None

    def parse_bool(value):
        raw = str(value or "").strip()
        if raw in ["", "是", "true", "True", "1", "yes"]:
            return True
        if raw in ["否", "false", "False", "0", "no"]:
            return False
        return True

    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.worksheets[0]

        results = {"success": 0, "failed": 0, "errors": []}
        status_allowed = {"available", "on_job", "paused", "blacklisted", "inactive"}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            first_cell = str(row[0]).strip()
            if first_cell.endswith("*") or first_cell == "":
                continue

            try:
                real_name = first_cell
                phone = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                id_card = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                gender = parse_gender(row[3] if len(row) > 3 else "")
                age = int(row[4]) if len(row) > 4 and row[4] not in [None, ""] else None
                experience_years = int(row[5]) if len(row) > 5 and row[5] not in [None, ""] else None
                job_types = split_multi_value(row[6] if len(row) > 6 else "")
                skills = split_multi_value(row[7] if len(row) > 7 else "")
                address = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                service_areas = split_multi_value(row[9] if len(row) > 9 else "")
                wechat = str(row[10]).strip() if len(row) > 10 and row[10] else ""
                expected_salary = float(row[11]) if len(row) > 11 and row[11] not in [None, ""] else None
                current_status = str(row[12]).strip() if len(row) > 12 and row[12] else "available"
                is_available = parse_bool(row[13] if len(row) > 13 else "")
                introduction = str(row[14]).strip() if len(row) > 14 and row[14] else ""

                if not all([real_name, phone, id_card, gender, age is not None, experience_years is not None, address, introduction]) or not skills:
                    results["errors"].append(f"第{row_idx}行：姓名格式错误")
                    results["failed"] += 1
                    continue

                if current_status not in status_allowed:
                    results["errors"].append(f"第{row_idx}行：手机号格式错误")
                    results["failed"] += 1
                    continue

                if db.query(User).filter(User.phone == phone).first():
                    results["errors"].append(f"第{row_idx}行：姓名格式错误")
                    results["failed"] += 1
                    continue

                if db.query(WorkerProfile).filter(WorkerProfile.id_card == id_card).first() or db.query(User).filter(User.id_card == id_card).first():
                    results["errors"].append(f"第{row_idx}行：手机号格式错误")
                    results["failed"] += 1
                    continue

                login_email = f"{id_card}@worker.local"
                while db.query(User).filter(User.email == login_email).first():
                    login_email = f"{id_card}_{generate_uuid()[:6]}@worker.local"

                user = User(
                    id=generate_uuid(),
                    email=login_email,
                    id_card=id_card,
                    phone=phone,
                    password_hash=get_password_hash(generate_uuid()),
                    nickname=real_name,
                    real_name=real_name,
                    role="worker",
                    status="active"
                )
                db.add(user)
                db.flush()

                worker = WorkerProfile(
                    id=generate_uuid(),
                    user_id=user.id,
                    recorder_staff_id=current_user_id,
                    real_name=real_name,
                    phone=phone,
                    id_card=id_card,
                    gender=gender,
                    age=age,
                    wechat=wechat or None,
                    address=address,
                    experience_years=experience_years,
                    skills=skills,
                    job_types=job_types,
                    service_areas=service_areas,
                    expected_salary=expected_salary,
                    introduction=introduction,
                    current_status=current_status,
                    is_available=is_available,
                    is_recommended=False,
                    rating=5.0,
                    total_orders=0,
                    completed_orders=0
                )
                db.add(worker)
                results["success"] += 1
            except Exception as e:
                results["errors"].append(f"第{row_idx}行：{str(e)}")
                results["failed"] += 1

        db.commit()
        return ApiResponse.success(
            data=results,
            message=f"导入完成：成功{results['success']}条，失败{results['failed']}条"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入失败: {str(e)}"
        )


@router.post("/import/users", summary="批量导入用户")
async def import_users(
    file: UploadFile = File(..., description="Excel文件"),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """
    批量导入用户
    
    - 上传按模板填写的Excel文件
    - 自动创建用户账号
    - 默认密码：123456
    """
    if user_role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传Excel文件")
    
    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        # 获取第一个sheet（数据sheet）
        ws = wb.worksheets[0]
        
        results = {"success": 0, "failed": 0, "errors": []}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not row[0]:
                continue
            
            first_cell = str(row[0]).strip()
            # 跳过说明行和表头行
            if first_cell.startswith("【") or first_cell.endswith("*") or first_cell == "邮箱":
                continue
            
            try:
                email = first_cell
                phone = str(row[1]).strip() if len(row) > 1 and row[1] else None
                nickname = str(row[2]).strip() if len(row) > 2 and row[2] else None
                
                if not email or '@' not in email:
                    results["errors"].append(f"第{row_idx}行: 邮箱格式不正确")
                    results["failed"] += 1
                    continue
                
                if db.query(User).filter(User.email == email).first():
                    results["errors"].append(f"第{row_idx}行: 邮箱{email}已存在")
                    results["failed"] += 1
                    continue
                
                if phone and db.query(User).filter(User.phone == phone).first():
                    results["errors"].append(f"第{row_idx}行: 手机号{phone}已存在")
                    results["failed"] += 1
                    continue
                
                user = User(
                    id=generate_uuid(),
                    email=email,
                    phone=phone,
                    password_hash=get_password_hash("123456"),
                    nickname=nickname,
                    role="user",
                    status="active"
                )
                db.add(user)
                results["success"] += 1
                
            except Exception as e:
                results["errors"].append(f"第{row_idx}行: {str(e)}")
                results["failed"] += 1
        
        db.commit()
        
        return ApiResponse.success(
            data=results,
            message=f"导入完成：成功{results['success']}条，失败{results['failed']}条"
        )
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入失败: {str(e)}"
        )


@router.post("/import/leads", summary="批量导入客户线索")
async def import_leads(
    file: UploadFile = File(..., description="Excel文件"),
    current_user_id: str = Depends(get_current_user_id),
    user_role: str = Depends(get_current_user_role),
    db: Session = Depends(get_db)
):
    """批量导入客户线索（管理员/员工）"""
    if user_role not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权操作")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传Excel文件")

    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.worksheets[0]

        results = {"success": 0, "failed": 0, "errors": []}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:
                continue

            first_cell = str(row[0]).strip() if row[0] else ""
            if first_cell.startswith("【") or first_cell.endswith("*") or first_cell == "线索日期":
                continue

            try:
                lead_date = parse_date_safe(str(row[0]).strip()) if row[0] else datetime.now().date()
                customer_name = str(row[1]).strip() if row[1] else ""
                phone = str(row[2]).strip() if row[2] else ""
                lead_category = str(row[3]).strip().upper() if row[3] else "B"
                service_type = str(row[4]).strip() if row[4] else ""
                source = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                source_detail = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                lead_name = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                demand_address = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                demand_detail = str(row[9]).strip() if len(row) > 9 and row[9] else ""
                budget_raw = row[10] if len(row) > 10 else None
                remark = str(row[11]).strip() if len(row) > 11 and row[11] else ""
                owner_staff_id_excel = str(row[12]).strip() if len(row) > 12 and row[12] else ""

                if not customer_name or not phone or not service_type:
                    results["errors"].append(f"第{row_idx}行: 缺少必填字段（客户称呼/联系电话/订单类型）")
                    results["failed"] += 1
                    continue

                if lead_category not in ["A", "B"]:
                    results["errors"].append(f"第{row_idx}行: 线索类别仅支持 A/B")
                    results["failed"] += 1
                    continue

                owner_staff_id = current_user_id
                if user_role == "admin" and owner_staff_id_excel:
                    owner_staff = db.query(User).filter(
                        User.id == owner_staff_id_excel, User.role == "staff", User.status == "active"
                    ).first()
                    if owner_staff:
                        owner_staff_id = owner_staff.id

                budget = None
                if budget_raw not in [None, ""]:
                    try:
                        budget = float(budget_raw)
                    except Exception:
                        results["errors"].append(f"第{row_idx}行: 预算格式不正确")
                        results["failed"] += 1
                        continue

                lead = CustomerLead(
                    id=generate_uuid(),
                    lead_no=f"LEAD{datetime.now().strftime('%Y%m%d%H%M%S')}{str(row_idx).zfill(3)}",
                    owner_staff_id=owner_staff_id,
                    source=source or None,
                    source_detail=source_detail or None,
                    lead_name=lead_name or None,
                    lead_type=None,
                    lead_date=lead_date,
                    customer_name=customer_name,
                    phone=phone,
                    service_type=service_type,
                    demand_address=demand_address or None,
                    demand_detail=demand_detail or None,
                    budget=budget,
                    lead_category=lead_category,
                    status="new",
                    remark=remark or None
                )
                db.add(lead)
                results["success"] += 1
            except Exception as e:
                results["errors"].append(f"第{row_idx}行: {str(e)}")
                results["failed"] += 1

        db.commit()
        return ApiResponse.success(
            data=results,
            message=f"导入完成：成功{results['success']}条，失败{results['failed']}条"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"导入失败: {str(e)}"
        )
